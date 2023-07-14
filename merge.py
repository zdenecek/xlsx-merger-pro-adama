
import csv

# the openpyxl library is used to read excel files
from openpyxl import load_workbook, Workbook

class Merger:
       
    def __init__(self, logger, save_extra_cell_to_column = []):
        """
        Args:
            save_extra_cell_to_column (list, optional): array of cell adresses to save to the output file (as extra columns). Defaults to [].
        """
        
        self.logger = logger
        self.save_extra_cell_to_column = save_extra_cell_to_column
    

    def merge_files_into_csv(self, file_list, output_file, lines_to_skip_in_each_file):

        with open(output_file, 'w', newline='') as outfile:                       
            writer = csv.writer(outfile)
            res = self.merge_files(file_list, writer.writerow, lines_to_skip_in_each_file)

        return res

    def merge_files_into_xlsx(self, file_list, output_file, lines_to_skip_in_each_file):

        workbook  = Workbook(write_only=True)
        sheet = workbook.create_sheet()

        res = self.merge_files(file_list, sheet.append, lines_to_skip_in_each_file)
        
        workbook.save(output_file)
        
        return res
                                            

    def merge_files(self, file_list, write_function, lines_to_skip):
        row_counter = 0
        for file in file_list:  
                self.logger.debug(f"Processing {file}")                                  
                workbook = load_workbook(filename=file, read_only=True)               
                sheet = workbook.worksheets[0]                 
                res = self.process_sheet(sheet, write_function, lines_to_skip)       
                row_counter += res
                    
                workbook.close()   
                self.logger.debug(f"Processed {res} lines from {file}")
                
        
        return row_counter

    def process_sheet(self, sheet, write_function, lines_to_skip):
        row_counter = 0
        extra_values = tuple(map(lambda x: self.get_cell_value(sheet, x), self.save_extra_cell_to_column))
        
        if len(extra_values) > 0:
            self.logger.debug(f"Found value for cells {','.join(self.save_extra_cell_to_column)}: {','.join(map(str,extra_values)) if len(extra_values) > 0 else 'None'}")
        
        for row_values in sheet.iter_rows(min_row = lines_to_skip + 1, values_only=True):      
            if(row_values[0] == None):
                continue
            write_function(row_values + extra_values)
            row_counter += 1
            
            
        return row_counter
    
    def get_cell_value(self, sheet, cell):
        return sheet[cell].value


if __name__ == "__main__":
    
    import sys
    
    def usage():
        print("Usage: python merge_files.py <lines_to_skip> <output_file> <input_file_1> <input_file_2> ...")
    
    if(len(sys.argv) < 4):
        usage()
        exit(1)
       
    try:
        lines = int(sys.argv[1])
    except:
        usage()
        exit(1)
        
    out = sys.argv[2]
    in_ = sys.argv[3:]
    
    merger = Merger()
    
    try:
        if out.endswith(".csv"):
            res =  merger.merge_files_into_csv(in_, out, lines)
        elif out.endswith(".xlsx"):
            res =  merger.merge_files_into_xlsx(in_, out, lines)
        else:
            res = merger.merge_files_into_csv(in_, out + ".csv", lines)
            
        print(f"Successfully saved {res} lines" )
    except Exception as e:
        print(f"Failed: " + str(e) )
    
    